"""
Script para procesar base predial Salgar 2026
- Reestructura 'archivo salgar convertido.xlsx' (5 secciones -> tabla plana)
- Actualiza avalúos urbanos en 'Base de datos predial 2024-2025_MAURO.xlsx'
- Calcula liquidaciones 2026 con límites Ley 44 de 1990
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

RUTA = "g:/Mi unidad/LEGAL/MUNICIPIOS/SALGAR/nuevo analisis 2026/"

# ===========================================================================
# 1. LEER Y PROCESAR ARCHIVO SALGAR CONVERTIDO
# ===========================================================================
print("Leyendo archivo salgar convertido.xlsx...")
wb_salgar = openpyxl.load_workbook(RUTA + "archivo salgar convertido.xlsx", data_only=True)
ws_salgar = wb_salgar[wb_salgar.sheetnames[0]]

# Mapeo de columnas por sección (índice base 0)
# Sección 1: predios
S1_PID   = 3   # predio_id
S1_ANO   = 1   # año catastral (2025)
S1_DIR   = 24  # dirección
S1_VIG   = 28  # vigencia resolución (ej: 1012026)
S1_RES   = 29  # código resolución
S1_TIPO  = 30  # tipo actuación
S1_CBML  = 32  # CBML 30 dígitos

# Sección 2: avalúos
S2_PID   = 1   # predio_id
S2_ANO   = 2   # año avalúo (2026)
S2_AT    = 3   # área terreno (dm²)
S2_AC    = 6   # área construcción (dm²)
S2_VT    = 10  # avalúo terreno
S2_AVTO  = 11  # avalúo total 2026

# Sección 3: propietarios
S3_PID   = 1   # predio_id
S3_DOC   = 4   # documento
S3_TDOC  = 5   # tipo documento (1=CC, 3=NIT...)
S3_APE1  = 6   # apellido 1
S3_APE2  = 7   # apellido 2
S3_NOM   = 8   # nombre(s)
S3_ENT   = 9   # entidad/razón social

# Sección 4: terrenos
S4_PID   = 1
S4_TIPO  = 2   # tipo terreno
S4_COEF  = 3   # coeficiente (%)

# Sección 5: construcciones
S5_PID   = 1
S5_NUM   = 2   # número construcción
S5_AREA  = 3   # área construcción (dm²)
S5_PISO  = 4   # piso/clasificación
S5_PTS   = 5   # puntos catastral
S5_USO   = 6   # uso/destino
S5_DEST  = 7   # sector económico (R, C, I, etc.)

print("Separando secciones...")
sec1, sec2, sec3, sec4, sec5 = {}, {}, {}, {}, {}

for row in ws_salgar.iter_rows(min_row=2, values_only=True):
    tipo = row[0]
    if tipo not in [1, 2, 3, 4, 5]:
        continue

    pid = int(row[S1_PID]) if tipo == 1 else int(row[1])

    if tipo == 1:
        cbml = str(row[S1_CBML]) if row[S1_CBML] else ""
        zona = cbml[5:7] if len(cbml) >= 7 else "XX"
        sector_map = {"01": "URBANO", "02": "RURAL", "03": "RURAL",
                      "04": "RURAL", "05": "RURAL"}
        sector = sector_map.get(zona, "RURAL")
        sec1[pid] = {
            "pid": pid,
            "ano_catastral": row[S1_ANO],
            "direccion": str(row[S1_DIR]).strip() if row[S1_DIR] else "",
            "vigencia": row[S1_VIG],
            "resolucion": str(row[S1_RES]).strip() if row[S1_RES] else "",
            "tipo_actuacion": str(row[S1_TIPO]).strip() if row[S1_TIPO] else "",
            "cbml": cbml,
            "zona": zona,
            "sector": sector,
        }

    elif tipo == 2:
        sec2[pid] = {
            "ano_avaluo": row[S2_ANO],
            "area_terreno_dm2": row[S2_AT],
            "area_const_dm2": row[S2_AC],
            "avaluo_terreno": row[S2_VT] or 0,
            "avaluo_total_2026": row[S2_AVTO] or 0,
        }

    elif tipo == 3:
        if pid not in sec3:  # tomar primer propietario
            nombre_parts = [
                str(row[S3_APE1]).strip() if row[S3_APE1] else "",
                str(row[S3_APE2]).strip() if row[S3_APE2] else "",
                str(row[S3_NOM]).strip() if row[S3_NOM] else "",
            ]
            entidad = str(row[S3_ENT]).strip() if row[S3_ENT] else ""
            nombre = " ".join(p for p in nombre_parts if p).strip()
            if not nombre and entidad:
                nombre = entidad
            doc = row[S3_DOC]
            sec3[pid] = {
                "documento": str(int(doc)) if isinstance(doc, float) else str(doc) if doc else "",
                "tipo_doc": row[S3_TDOC],
                "nombre_propietario": nombre or entidad,
            }

    elif tipo == 4:
        if pid not in sec4:
            sec4[pid] = {"tipo_terreno": row[S4_TIPO], "coef_terreno": row[S4_COEF]}

    elif tipo == 5:
        if pid not in sec5:  # primera construcción
            sec5[pid] = {
                "area_const_dm2_s5": row[S5_AREA],
                "piso": row[S5_PISO],
                "puntos": row[S5_PTS],
                "uso_destino": str(row[S5_USO]).strip() if row[S5_USO] else "",
                "sector_eco": str(row[S5_DEST]).strip() if row[S5_DEST] else "",
            }

print(f"  Sección 1 (predios):       {len(sec1):,} registros")
print(f"  Sección 2 (avalúos 2026):  {len(sec2):,} registros")
print(f"  Sección 3 (propietarios):  {len(sec3):,} registros")
print(f"  Sección 4 (terrenos):      {len(sec4):,} registros")
print(f"  Sección 5 (construcciones):{len(sec5):,} registros")

# ===========================================================================
# 2. LEER BASE MAURO 2024-2025
# ===========================================================================
print("\nLeyendo Base de datos predial 2024-2025_MAURO.xlsx...")
wb_mauro = openpyxl.load_workbook(RUTA + "Base de datos predial 2024-2025_MAURO.xlsx", data_only=True)
ws_mauro = wb_mauro["Hoja1"]

mauro_data = {}
for row in ws_mauro.iter_rows(min_row=2, values_only=True):
    ficha_raw = row[2]
    try:
        ficha = int(float(str(ficha_raw)))
    except:
        continue
    mauro_data[ficha] = {
        "documento": str(row[0]) if row[0] else "",
        "tercero": str(row[1]) if row[1] else "",
        "ficha": ficha,
        "es_exento": row[3],
        "avaluo_ant": row[4] or 0,
        "tarifa_ant": row[5] or 0,
        "liq_2024": row[6] or 0,
        "estrato_ant": row[7] or 0,
        "dest_ant": str(row[8]) if row[8] else "",
        "sector_ant": str(row[9]) if row[9] else "",
        "area_ant": row[10] or 0,
        "area_const_ant": row[11] or 0,
        "avaluo_act": row[12] or 0,   # avaluo 2025
        "tarifa_act": row[13] or 0,
        "liq_2025": row[14] or 0,
        "estrato_act": row[15] or 0,
        "dest_act": str(row[16]) if row[16] else "",
        "sector_act": str(row[17]) if row[17] else "",
        "area_act": row[18] or 0,
        "area_const_act": row[19] or 0,
        "periodo_pagado": row[20],
        "deuda": row[21] or 0,
    }

print(f"  Total predios en MAURO: {len(mauro_data):,}")

# ===========================================================================
# 3. CREAR LIBRO DE SALIDA
# ===========================================================================
print("\nCreando archivo de salida...")
wb_out = openpyxl.Workbook()

# Estilos
FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_HEADER2 = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
FILL_ALERTA = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
FILL_LIMITE = PatternFill(start_color="FF7C80", end_color="FF7C80", fill_type="solid")
FILL_OK = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
FILL_NUEVO = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FONT_H = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_N = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center")
ALIGN_R = Alignment(horizontal="right", vertical="center")


def style_header(cell, fill=FILL_HEADER):
    cell.fill = fill
    cell.font = FONT_H
    cell.alignment = ALIGN_C


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ===========================================================================
# HOJA 1: TABLA PLANA SALGAR CONVERTIDO
# ===========================================================================
print("  Generando hoja: Tabla_Salgar...")
ws_tab = wb_out.active
ws_tab.title = "Tabla_Salgar"
ws_tab.freeze_panes = "A2"

headers_tab = [
    "PREDIO_ID", "AÑO_CATASTRAL", "SECTOR", "ZONA", "CBML",
    "DIRECCIÓN", "VIGENCIA_RES", "RESOLUCIÓN", "TIPO_ACTUACIÓN",
    "DOCUMENTO", "TIPO_DOC", "NOMBRE_PROPIETARIO",
    "ÁREA_TERRENO_dm2", "ÁREA_CONST_dm2", "AVALÚO_TERRENO_2026", "AVALÚO_CONST_2026", "AVALÚO_TOTAL_2026",
    "USO_DESTINO", "SECTOR_ECO", "PISO", "PUNTOS",
    "TIPO_TERRENO", "COEF_TERRENO_%",
    # Datos MAURO para comparación
    "TERCERO_MAURO", "AVALÚO_2025_MAURO", "TARIFA_ACT", "LIQ_2025", "DEST_MAURO", "EN_MAURO",
]

for c, h in enumerate(headers_tab, 1):
    cell = ws_tab.cell(row=1, column=c, value=h)
    style_header(cell)

ws_tab.row_dimensions[1].height = 35

row_num = 2
for pid, s1 in sorted(sec1.items()):
    s2 = sec2.get(pid, {})
    s3 = sec3.get(pid, {})
    s4 = sec4.get(pid, {})
    s5 = sec5.get(pid, {})
    m = mauro_data.get(pid, {})

    avaluo_total = s2.get("avaluo_total_2026", 0)
    avaluo_terreno = s2.get("avaluo_terreno", 0)
    avaluo_const = (avaluo_total - avaluo_terreno) if avaluo_total and avaluo_terreno else 0

    vals = [
        pid,
        s1.get("ano_catastral"),
        s1.get("sector"),
        s1.get("zona"),
        s1.get("cbml"),
        s1.get("direccion"),
        s1.get("vigencia"),
        s1.get("resolucion"),
        s1.get("tipo_actuacion"),
        s3.get("documento", ""),
        s3.get("tipo_doc", ""),
        s3.get("nombre_propietario", ""),
        s2.get("area_terreno_dm2"),
        s2.get("area_const_dm2"),
        avaluo_terreno,
        avaluo_const,
        avaluo_total,
        s5.get("uso_destino", ""),
        s5.get("sector_eco", ""),
        s5.get("piso"),
        s5.get("puntos"),
        s4.get("tipo_terreno"),
        s4.get("coef_terreno"),
        m.get("tercero", ""),
        m.get("avaluo_act", ""),
        m.get("tarifa_act", ""),
        m.get("liq_2025", ""),
        m.get("dest_act", ""),
        "SI" if m else "NO",
    ]

    for c, v in enumerate(vals, 1):
        cell = ws_tab.cell(row=row_num, column=c, value=v)
        cell.font = FONT_N
        cell.alignment = ALIGN_L
        # Formato numérico para avalúos
        if c in [15, 16, 17, 25]:
            cell.alignment = ALIGN_R
            if isinstance(v, (int, float)) and v:
                cell.number_format = '#,##0'
        if c in [13, 14]:
            cell.alignment = ALIGN_R

    # Color por sector
    if s1.get("sector") == "URBANO":
        for c in range(1, len(headers_tab) + 1):
            ws_tab.cell(row=row_num, column=c).fill = PatternFill(
                start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

    row_num += 1

set_col_widths(ws_tab, {
    "A": 12, "B": 10, "C": 10, "D": 6, "E": 32, "F": 35, "G": 12,
    "H": 14, "I": 22, "J": 14, "K": 8, "L": 35, "M": 14, "N": 14,
    "O": 18, "P": 18, "Q": 18, "R": 40, "S": 10, "T": 6, "U": 8,
    "V": 12, "W": 12, "X": 35, "Y": 16, "Z": 10, "AA": 14, "AB": 20, "AC": 8,
})
print(f"    Total filas: {row_num - 2:,}")

# ===========================================================================
# HOJA 2: PREDIOS URBANOS 2026 + LEY 44
# ===========================================================================
print("  Generando hoja: Predios_Urbanos_2026...")
ws_urb = wb_out.create_sheet("Predios_Urbanos_2026")
ws_urb.freeze_panes = "A2"

headers_urb = [
    # Identificación
    "FICHA", "DOCUMENTO", "TERCERO", "DIRECCIÓN_IGAC", "ES_EXENTO",
    # Situación 2025
    "AVALÚO_2025", "TARIFA_2025", "LIQ_2025", "DEST_2025", "ÁREA_CONST_m2",
    # Nuevos avalúos 2026 (IGAC)
    "AVALÚO_TERRENO_2026", "AVALÚO_CONST_2026", "AVALÚO_TOTAL_2026",
    "VARIACIÓN_AVALÚO_%", "TIPO_ACTUACIÓN",
    # Uso/destino nuevo
    "USO_DESTINO_IGAC", "SECTOR_ECO",
    # Cálculo Ley 44
    "TARIFA_APLICADA", "LIQ_2026_SIN_LÍMITE",
    "LÍMITE_LEY44_2x", "LIQ_2026_CON_LÍMITE",
    "APLICA_LÍMITE", "AHORRO_CONTRIBUYENTE",
    "INCREMENTO_REAL_$", "INCREMENTO_REAL_%",
    # Estado
    "EN_SALGAR_IGAC", "NOVEDADES",
]

for c, h in enumerate(headers_urb, 1):
    cell = ws_urb.cell(row=1, column=c, value=h)
    style_header(cell, FILL_HEADER2)

ws_urb.row_dimensions[1].height = 40

# Grupos de columnas con color para segunda fila de encabezado
grupos = {
    "IDENTIFICACIÓN": (1, 5),
    "SITUACIÓN 2025": (6, 10),
    "NUEVOS AVALÚOS 2026 (IGAC - Formación y Actualización)": (11, 15),
    "USO/DESTINO": (16, 17),
    "LIQUIDACIÓN 2026 - LEY 44/1990": (18, 25),
    "NOVEDADES": (26, 27),
}

# Filas de datos
row_num = 2
stats = {
    "total_urbanos_mauro": 0,
    "con_igac": 0,
    "sin_igac": 0,
    "aplica_limite": 0,
    "no_aplica": 0,
    "exentos": 0,
    "suma_liq_2025": 0,
    "suma_liq_2026_sin": 0,
    "suma_liq_2026_con": 0,
    "nuevos_predios_igac": 0,
}

# 1. Predios MAURO urbanos con datos IGAC
for ficha, m in sorted(mauro_data.items(), key=lambda x: str(x[0])):
    if m.get("sector_act") != "URBANO" and m.get("sector_ant") != "URBANO":
        continue

    stats["total_urbanos_mauro"] += 1

    s1 = sec1.get(ficha, {})
    s2 = sec2.get(ficha, {})
    s3 = sec3.get(ficha, {})
    s5 = sec5.get(ficha, {})

    tiene_igac = bool(s2)

    avaluo_2025 = m.get("avaluo_act", 0)
    liq_2025 = m.get("liq_2025", 0)
    tarifa_act = m.get("tarifa_act", 0)
    es_exento = m.get("es_exento", "N")

    # Tarifa efectiva (para predios con tarifa_act=0 pero liq_2025>0)
    if tarifa_act == 0 and liq_2025 > 0 and avaluo_2025 > 0:
        tarifa_efectiva = round(liq_2025 / avaluo_2025 * 1000, 2)
    else:
        tarifa_efectiva = tarifa_act

    avaluo_2026 = s2.get("avaluo_total_2026", 0) if tiene_igac else 0
    avaluo_terreno_2026 = s2.get("avaluo_terreno", 0) if tiene_igac else 0
    avaluo_const_2026 = (avaluo_2026 - avaluo_terreno_2026) if avaluo_2026 else 0

    # Variación %
    var_pct = None
    if avaluo_2025 > 0 and avaluo_2026 > 0:
        var_pct = round((avaluo_2026 - avaluo_2025) / avaluo_2025 * 100, 2)

    # Liquidación 2026 sin límite
    liq_2026_sin = 0
    if avaluo_2026 > 0 and tarifa_efectiva > 0:
        liq_2026_sin = round(avaluo_2026 * tarifa_efectiva / 1000)

    # Límite Ley 44: máximo el doble del año anterior
    limite_ley44 = round(liq_2025 * 2) if liq_2025 > 0 else 0

    # Liquidación final
    aplica_limite = False
    novedades = []
    if es_exento == "S":
        liq_2026_con = 0
        stats["exentos"] += 1
        novedades.append("EXENTO")
    elif liq_2025 == 0 or liq_2025 is None:
        liq_2026_con = liq_2026_sin
        novedades.append("Sin liquidación 2025")
    elif avaluo_2026 == 0:
        liq_2026_con = liq_2025  # Sin datos IGAC, mantiene 2025
        novedades.append("Sin datos IGAC 2026")
    else:
        if liq_2026_sin > limite_ley44 and limite_ley44 > 0:
            liq_2026_con = limite_ley44
            aplica_limite = True
            stats["aplica_limite"] += 1
        else:
            liq_2026_con = liq_2026_sin
            stats["no_aplica"] += 1

    ahorro = max(0, liq_2026_sin - liq_2026_con)
    incremento_real = liq_2026_con - liq_2025
    incremento_pct = round(incremento_real / liq_2025 * 100, 2) if liq_2025 > 0 else None

    if tiene_igac:
        stats["con_igac"] += 1
    else:
        stats["sin_igac"] += 1

    stats["suma_liq_2025"] += liq_2025 or 0
    stats["suma_liq_2026_sin"] += liq_2026_sin
    stats["suma_liq_2026_con"] += liq_2026_con

    if not novedades and not tiene_igac:
        novedades.append("Predio sin formación IGAC")

    vals = [
        ficha,
        m.get("documento"),
        m.get("tercero"),
        s1.get("direccion", ""),
        es_exento,
        # 2025
        avaluo_2025,
        tarifa_efectiva,
        liq_2025,
        m.get("dest_act"),
        m.get("area_const_act"),
        # 2026
        avaluo_terreno_2026 or None,
        avaluo_const_2026 or None,
        avaluo_2026 or None,
        var_pct,
        s1.get("tipo_actuacion", ""),
        # uso
        s5.get("uso_destino", ""),
        s5.get("sector_eco", ""),
        # ley 44
        tarifa_efectiva,
        liq_2026_sin or None,
        limite_ley44 or None,
        liq_2026_con or None,
        "SÍ" if aplica_limite else "NO",
        ahorro or None,
        incremento_real,
        incremento_pct,
        # estado
        "SÍ" if tiene_igac else "NO",
        "; ".join(novedades) if novedades else "",
    ]

    COLS_PESOS = {6, 8, 11, 12, 13, 19, 20, 21, 23, 24}
    COLS_PCT = {7, 14, 18, 25}

    for c, v in enumerate(vals, 1):
        cell = ws_urb.cell(row=row_num, column=c, value=v)
        cell.font = FONT_N
        cell.alignment = ALIGN_L
        if c in COLS_PESOS:
            cell.alignment = ALIGN_R
            if isinstance(v, (int, float)) and v:
                cell.number_format = '#,##0'
        elif c in COLS_PCT:
            cell.alignment = ALIGN_R
            if isinstance(v, (int, float)):
                cell.number_format = '#,##0.00'

    # Colores por estado
    if aplica_limite:
        ws_urb.cell(row=row_num, column=22).fill = FILL_LIMITE
    elif not tiene_igac:
        ws_urb.cell(row=row_num, column=26).fill = FILL_ALERTA
    else:
        ws_urb.cell(row=row_num, column=22).fill = FILL_OK

    row_num += 1

# 2. Predios NUEVOS en IGAC (no están en MAURO)
for pid, s1 in sorted(sec1.items()):
    if s1.get("sector") != "URBANO":
        continue
    if pid in mauro_data:
        continue

    # Predio nuevo de la formación
    stats["nuevos_predios_igac"] += 1
    s2 = sec2.get(pid, {})
    s3 = sec3.get(pid, {})
    s5 = sec5.get(pid, {})

    avaluo_2026 = s2.get("avaluo_total_2026", 0)
    avaluo_terreno_2026 = s2.get("avaluo_terreno", 0)
    avaluo_const_2026 = (avaluo_2026 - avaluo_terreno_2026) if avaluo_2026 else 0

    vals = [
        pid,
        s3.get("documento", ""),
        s3.get("nombre_propietario", ""),
        s1.get("direccion", ""),
        "N",
        0, 0, 0, "", 0,
        avaluo_terreno_2026 or None,
        avaluo_const_2026 or None,
        avaluo_2026 or None,
        None, s1.get("tipo_actuacion", ""),
        s5.get("uso_destino", ""),
        s5.get("sector_eco", ""),
        0, 0, 0, 0,
        "N/A", 0, 0, None,
        "SÍ",
        "PREDIO NUEVO - Sin liquidación anterior",
    ]

    COLS_PESOS = {6, 8, 11, 12, 13, 19, 20, 21, 23, 24}

    for c, v in enumerate(vals, 1):
        cell = ws_urb.cell(row=row_num, column=c, value=v)
        cell.font = FONT_N
        cell.alignment = ALIGN_L
        if c in COLS_PESOS and isinstance(v, (int, float)) and v:
            cell.alignment = ALIGN_R
            cell.number_format = '#,##0'

    for c in range(1, len(headers_urb) + 1):
        ws_urb.cell(row=row_num, column=c).fill = FILL_NUEVO

    row_num += 1

set_col_widths(ws_urb, {
    "A": 12, "B": 13, "C": 38, "D": 35, "E": 9,
    "F": 16, "G": 10, "H": 14, "I": 16, "J": 11,
    "K": 16, "L": 16, "M": 18, "N": 14, "O": 24,
    "P": 38, "Q": 10,
    "R": 14, "S": 18, "T": 16, "U": 18, "V": 12, "W": 18,
    "X": 18, "Y": 13,
    "Z": 12, "AA": 32,
})

print(f"    Total filas urbanas: {row_num - 2:,}")

# ===========================================================================
# HOJA 3: RESUMEN ESTADÍSTICO
# ===========================================================================
print("  Generando hoja: Resumen_Ley44...")
ws_res = wb_out.create_sheet("Resumen_Ley44")

titulo_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
sub_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

def write_title(ws, row, text, fill=titulo_fill, col=1, colspan=3):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)
    ws.row_dimensions[row].height = 25
    return cell

def write_row(ws, row, label, value, fmt="#,##0", col=1):
    c1 = ws.cell(row=row, column=col, value=label)
    c2 = ws.cell(row=row, column=col + 1, value=value)
    c1.font = Font(name="Calibri", bold=True, size=10)
    c2.font = Font(name="Calibri", size=10)
    c2.alignment = Alignment(horizontal="right")
    if fmt and isinstance(value, (int, float)):
        c2.number_format = fmt

r = 1
ws_res.column_dimensions["A"].width = 45
ws_res.column_dimensions["B"].width = 20
ws_res.column_dimensions["C"].width = 20
ws_res.column_dimensions["D"].width = 20

write_title(ws_res, r, "ANÁLISIS PREDIAL SALGAR - LEY 44/1990 - FORMACIÓN CATASTRAL 2026", colspan=4)
r += 1

ws_res.cell(r, 1, "Elaborado con base en: archivo salgar convertido.xlsx + Base de datos predial 2024-2025_MAURO.xlsx")
ws_res.cell(r, 1).font = Font(name="Calibri", italic=True, size=9)
ws_res.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 2

write_title(ws_res, r, "1. UNIVERSO DE PREDIOS", fill=PatternFill("solid", fgColor="2E75B6"), colspan=4)
r += 1
write_row(ws_res, r, "Total predios en MAURO 2025 (urbanos)", stats["total_urbanos_mauro"]); r += 1
write_row(ws_res, r, "  Con datos IGAC 2026 (formación/actualización)", stats["con_igac"]); r += 1
write_row(ws_res, r, "  Sin datos IGAC 2026 (sin novedad de formación)", stats["sin_igac"]); r += 1
write_row(ws_res, r, "Predios NUEVOS en IGAC (no en MAURO)", stats["nuevos_predios_igac"]); r += 1
write_row(ws_res, r, "Total predios urbanos sección 1 IGAC", len([p for p,s in sec1.items() if s["sector"]=="URBANO"])); r += 2

write_title(ws_res, r, "2. IMPACTO FISCAL 2025 vs 2026", fill=PatternFill("solid", fgColor="2E75B6"), colspan=4)
r += 1
write_row(ws_res, r, "Recaudo actual 2025 ($)", stats["suma_liq_2025"]); r += 1
write_row(ws_res, r, "Recaudo bruto 2026 (sin límite Ley 44) ($)", stats["suma_liq_2026_sin"]); r += 1
write_row(ws_res, r, "Recaudo proyectado 2026 (con límite Ley 44) ($)", stats["suma_liq_2026_con"]); r += 1
if stats["suma_liq_2025"] > 0:
    var_recaudo = (stats["suma_liq_2026_con"] - stats["suma_liq_2025"]) / stats["suma_liq_2025"] * 100
    write_row(ws_res, r, "Variación recaudo 2025→2026 con Ley 44 (%)", round(var_recaudo, 2), fmt="#,##0.00"); r += 1
    benef_total = stats["suma_liq_2026_sin"] - stats["suma_liq_2026_con"]
    write_row(ws_res, r, "Beneficio total contribuyentes por Ley 44 ($)", benef_total); r += 2
else:
    r += 2

write_title(ws_res, r, "3. APLICACIÓN LÍMITE LEY 44 / 1990", fill=PatternFill("solid", fgColor="2E75B6"), colspan=4)
r += 1
write_row(ws_res, r, "Predios donde aplica el límite 2x Ley 44", stats["aplica_limite"]); r += 1
write_row(ws_res, r, "Predios donde NO aplica límite (2026 < 2x LIQ_2025)", stats["no_aplica"]); r += 1
write_row(ws_res, r, "Predios exentos", stats["exentos"]); r += 2

write_title(ws_res, r, "4. REFERENCIA NORMATIVA", fill=PatternFill("solid", fgColor="2E75B6"), colspan=4)
r += 1
normas = [
    ("Ley 44 de 1990 Art. 6", "Límite: nuevo impuesto no puede exceder 2x el del año anterior"),
    ("Ley 1955 de 2019", "Catastro Multipropósito - formación/actualización catastral"),
    ("Ley 1819 de 2016 Art. 354", "Tarifas máximas predial unificado (1‰ a 16‰)"),
    ("Decreto 1983 de 2017", "Tasas para predios en proceso de formación"),
]
for norma, desc in normas:
    ws_res.cell(r, 1, norma).font = Font(name="Calibri", bold=True, size=9)
    ws_res.cell(r, 2, desc).font = Font(name="Calibri", size=9)
    ws_res.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

r += 1
ws_res.cell(r, 1, "NOTA: El límite de la Ley 44 aquí calculado = LIQ_2025 × 2 (máximo 100% de incremento). " \
    "Verificar con el estatuto tributario municipal vigente la tarifa para 2026.").font = Font(name="Calibri", italic=True, size=9, color="FF0000")
ws_res.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

# ===========================================================================
# GUARDAR
# ===========================================================================
salida = RUTA + "Predial_Salgar_2026_Ley44.xlsx"
print(f"\nGuardando: {salida}")
wb_out.save(salida)
print("\n=== RESUMEN FINAL ===")
print(f"Predios totales en base MAURO:     {len(mauro_data):,}")
print(f"Predios en IGAC sección 1:         {len(sec1):,}")
print(f"  - Urbanos:                       {len([p for p,s in sec1.items() if s['sector']=='URBANO']):,}")
print(f"  - Rurales:                       {len([p for p,s in sec1.items() if s['sector']!='URBANO']):,}")
print(f"\nPredios urbanos MAURO:             {stats['total_urbanos_mauro']:,}")
print(f"  Con datos IGAC 2026:             {stats['con_igac']:,}")
print(f"  Sin datos IGAC:                  {stats['sin_igac']:,}")
print(f"  Predios nuevos IGAC:             {stats['nuevos_predios_igac']:,}")
print(f"\nAplica límite Ley 44:             {stats['aplica_limite']:,} predios")
print(f"No aplica límite:                  {stats['no_aplica']:,} predios")
print(f"\nRecaudo 2025:                      ${stats['suma_liq_2025']:>20,.0f}")
print(f"Recaudo 2026 bruto (sin límite):   ${stats['suma_liq_2026_sin']:>20,.0f}")
print(f"Recaudo 2026 con límite Ley 44:    ${stats['suma_liq_2026_con']:>20,.0f}")
if stats["suma_liq_2025"] > 0:
    benef = stats['suma_liq_2026_sin'] - stats['suma_liq_2026_con']
    var = (stats['suma_liq_2026_con'] - stats['suma_liq_2025']) / stats['suma_liq_2025'] * 100
    print(f"Beneficio contribuyentes:          ${benef:>20,.0f}")
    print(f"Variación recaudo:                 {var:>20.1f}%")
print(f"\nArchivo generado: Predial_Salgar_2026_Ley44.xlsx")

